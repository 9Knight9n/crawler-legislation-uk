import os.path
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from extract import headers
from extract.act import get_act_details, get_links, get_txt
from extract.detector import detect_type
from save import files_dir, files_list_dir
from utils import fix_dir_name, trim

excel = pd.read_excel(files_list_dir,)
wb = load_workbook(files_list_dir)
ws = wb.worksheets[0]

# wb2 = load_workbook(ref_list_dir)
# ws2 = wb2.worksheets[0]


def dl_file(url,name):
    path = files_dir+"/"+name
    if os.path.isfile(files_dir+"/"+name):
        print(f'file {url} already exist as {path} .skipping...')
        return
    r = requests.get(url)
    open(files_dir+"/"+name, 'wb').write(r.content)




def append_act(p_id:str):
    temp = p_id.split("/")
    type_ = temp[0]
    year = temp[1]
    num = temp[2]
    if already_added(type_, year, num):
        print(f'Act {p_id} already loaded.')
        return True
    act_detail = get_act_details(p_id)
    title = fix_dir_name(act_detail['title'])
    type_ = detect_type(type_,title)
    if type_ is None:
        print(f'Act {p_id} is not  included in accepted types.')
        return None
    ws.append([trim(type_),trim(year),trim(num),trim(title),trim(act_detail['extend']),trim(act_detail['note'])])
    wb.save(files_list_dir)
    for key in act_detail['files'].keys():
        # dl_file(act_detail['files'][key],trim(title)+key)
        if "xht" in key:
            txt = get_txt(act_detail['files'][key])
            text_file = open(files_dir+"/"+trim(title) + key[:-3] + "txt", "w")
            text_file.write(txt)
            text_file.close()

    # refs = get_links(act_detail['files']['.xht'])
    # for ref in refs:
    #     add_links(p_id,ref)
    return True



def add_links(p_id,link):
    ws2.append([p_id, link])
    wb2.save(ref_list_dir)




def already_added(type_,year,num):
    t_index = -1
    y_index = -1
    n_index = -1
    for index , t in enumerate(excel['type']):
        if t==type_:
            t_index = index
    for index , t in enumerate(excel['year']):
        if str(t)==year:
            y_index = index
    for index , t in enumerate(excel['number']):
        if str(t)==num:
            n_index = index
    if t_index == y_index and t_index==n_index and t_index != -1:
        return True
    else:
        return False





# append_act("eudn/2020/2255")